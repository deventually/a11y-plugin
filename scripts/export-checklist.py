#!/usr/bin/env python3
"""Export accessibility_checklist_audit.xlsx to checklist/wcag22-checklist.json.

Reads the Master Checklist sheet and merges with enrichment data to produce
the full policy layer JSON. Re-runnable: always overwrites the output.

Usage:
    python3 scripts/export-checklist.py
"""

import json
import os
import re
import sys

import openpyxl


def slugify(text: str) -> str:
    """Convert text to a URL-friendly slug."""
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"[\s-]+", "-", text)
    return text.strip("-")


def make_id(category: str, item: str, counters: dict) -> str:
    """Generate a check ID like 'page-structure-001'."""
    cat_slug = slugify(category)
    words = slugify(item).split("-")[:3]
    key = f"{cat_slug}-{'-'.join(words)}"
    counters[key] = counters.get(key, 0) + 1
    return f"{key}-{counters[key]:03d}"


def parse_wcag_sc(sc_string: str) -> list:
    """Split '1.3.1, 2.4.6' into ['1.3.1', '2.4.6']."""
    if not sc_string or sc_string == "Non-WCAG":
        return []
    return [s.strip() for s in sc_string.split(",")]


def parse_applies_to(applies_to: str) -> list:
    """Normalize applies_to into a list."""
    if not applies_to:
        return ["all"]
    parts = re.split(r"\s*/\s*|\s*,\s*", applies_to.strip())
    return [p.strip().lower() for p in parts if p.strip()]


ENRICHMENTS = {
    ("Page structure", "Each page has a unique, descriptive titl"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Document has a non-empty <title> element",
        "fail_condition": "No <title> element or <title> is empty",
        "suggested_fix": "Add a descriptive <title> element in the <head>",
        "severity": "critical",
        "file_patterns": ["*.html"],
    },
    ("Page structure", "The page language is declared correctly"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "<html> element has a valid lang attribute with a BCP 47 language tag",
        "fail_condition": "<html> element has no lang attribute or an empty/invalid value",
        "suggested_fix": "Add lang attribute to <html>, e.g. <html lang=\"en\">",
        "severity": "critical",
        "file_patterns": ["*.html"],
    },
    ("Page structure", "Headings reflect the content structure i"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Heading levels (h1-h6) do not skip levels and h1 appears first",
        "fail_condition": "Heading levels skip (e.g., h1 to h3) or no h1 is present",
        "suggested_fix": "Restructure headings to follow a logical hierarchy without skipping levels",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Page structure", "Semantic HTML is used appropriately for "): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Navigation uses <nav>, main content uses <main>, buttons use <button>, links use <a>",
        "fail_condition": "Non-semantic elements (div/span) used for interactive or structural roles without ARIA",
        "suggested_fix": "Replace <div> with appropriate semantic elements (<nav>, <main>, <button>, <a>, etc.)",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Page structure", "Repeated content can be bypassed, such a"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "A skip link or bypass mechanism exists near the start of the document",
        "fail_condition": "No skip link, <main> landmark, or other bypass mechanism found",
        "suggested_fix": "Add a skip link as the first focusable element: <a href=\"#main\" class=\"skip-link\">Skip to content</a>",
        "severity": "critical",
        "file_patterns": ["*.html"],
    },
    ("Page structure", "Information and instructions are not con"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Instructions and information use text labels, not just color/shape/location/sound",
        "fail_condition": "Directions rely solely on sensory characteristics (e.g., 'click the red button')",
        "suggested_fix": "Add text labels alongside visual/auditory cues",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Page structure", "Repeated navigation and repeated UI appe"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Repeated navigation components appear in the same relative order across pages",
        "fail_condition": "Navigation order varies between pages without clear reason",
        "suggested_fix": "Ensure shared navigation components maintain consistent ordering across all pages",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Keyboard and focus", "All functionality works with a keyboard "): {
        "automatable": "partial",
        "inputs_required": ["html", "keyboard"],
        "pass_condition": "All interactive elements are natively focusable or have tabindex and keyboard event handlers",
        "fail_condition": "Interactive element uses only mouse events (onclick on div) without keyboard support",
        "suggested_fix": "Use native interactive elements (<button>, <a>) or add tabindex='0' and keydown/keypress handlers",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Keyboard and focus", "There are no keyboard traps"): {
        "automatable": "partial",
        "inputs_required": ["html", "keyboard"],
        "pass_condition": "No element prevents focus from moving away via Tab or Escape",
        "fail_condition": "Focus is trapped in an element with no documented escape mechanism",
        "suggested_fix": "Ensure all focus-trapping components (modals) provide keyboard escape, and non-modal components don't trap focus",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Keyboard and focus", "Focus order is logical"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "No positive tabindex values; DOM order matches visual reading order",
        "fail_condition": "Positive tabindex values force non-logical focus order, or DOM order mismatches visual layout",
        "suggested_fix": "Remove positive tabindex values and reorder DOM to match visual layout",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Keyboard and focus", "Focus moves appropriately when opening o"): {
        "automatable": "partial",
        "inputs_required": ["html", "keyboard"],
        "pass_condition": "Dynamic UI (dialogs, menus) moves focus on open and restores on close",
        "fail_condition": "Focus does not move when dynamic content appears or is not restored when it closes",
        "suggested_fix": "Add focus management: move focus to new content on open, return to trigger on close",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Keyboard and focus", "Focus is clearly visible"): {
        "automatable": "partial",
        "inputs_required": ["html", "dom"],
        "pass_condition": "No CSS rules that remove focus indicators (outline: none/0) without replacement",
        "fail_condition": "CSS removes focus outline without providing an alternative visible indicator",
        "suggested_fix": "Remove 'outline: none' or 'outline: 0' rules, or replace with a visible custom focus style",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Keyboard and focus", "Focused elements are not hidden behind s"): {
        "automatable": "partial",
        "inputs_required": ["html", "dom"],
        "pass_condition": "No fixed/sticky positioned elements that could obscure focused content without scroll-padding or equivalent",
        "fail_condition": "Fixed/sticky elements exist without scroll-margin/scroll-padding to prevent focus obscurance",
        "suggested_fix": "Add scroll-padding-top to account for sticky header height, or use scroll-margin on focusable elements",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Links, buttons, and controls", "Links are used for navigation and button"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "<a> elements have href and navigate; <button> elements trigger actions; no role misuse",
        "fail_condition": "<a> without href used as button, <button> used for navigation, or div/span used for either without proper role",
        "suggested_fix": "Use <a href> for navigation and <button> for actions. Replace div/span click handlers with proper elements",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Links, buttons, and controls", "Link purpose is clear from the text or i"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Links have descriptive text content, aria-label, or aria-labelledby; no bare 'click here' or 'read more'",
        "fail_condition": "Link text is generic ('click here', 'read more', 'here') without aria-label providing context",
        "suggested_fix": "Replace generic link text with descriptive text, or add aria-label describing the destination",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Links, buttons, and controls", "Icon-only controls have an accessible na"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "Every button/link containing only an icon has aria-label, aria-labelledby, or visually-hidden text",
        "fail_condition": "Button or link contains only an icon (<svg>, <i>, <img>) with no accessible name",
        "suggested_fix": "Add aria-label to the button/link, or add a visually-hidden <span> with descriptive text",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Links, buttons, and controls", "Custom controls expose the correct name,"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Custom interactive elements have appropriate role, aria-label/aria-labelledby, and state attributes",
        "fail_condition": "Custom component (div/span) acts as a control without role, accessible name, or state attributes",
        "suggested_fix": "Add appropriate ARIA role, accessible name (aria-label), and state attributes (aria-expanded, etc.)",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Links, buttons, and controls", "Controls do not trigger unexpected conte"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "No onfocus/onchange handlers that navigate or significantly alter the page without user activation",
        "fail_condition": "Form control or link triggers navigation or major UI change on focus or input without explicit submit",
        "suggested_fix": "Move context-changing behavior to explicit user actions (button click, form submit)",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Links, buttons, and controls", "Pointer targets are large enough except "): {
        "automatable": "partial",
        "inputs_required": ["html", "dom"],
        "pass_condition": "Interactive elements have explicit minimum dimensions of 24x24 CSS pixels or are inline/sentence text",
        "fail_condition": "Interactive element has explicit dimensions below 24x24 CSS pixels",
        "suggested_fix": "Increase target size to at least 24x24 CSS pixels using min-width/min-height or padding",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Color is not the only way information, e"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Error states, required fields, and status indicators use text/icons in addition to color",
        "fail_condition": "Information conveyed by color alone with no text, icon, or pattern alternative",
        "suggested_fix": "Add text labels, icons, or patterns alongside color-based indicators",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Normal text contrast is at least 4.5:1"): {
        "automatable": "partial",
        "inputs_required": ["html", "dom"],
        "pass_condition": "Hardcoded text colors meet 4.5:1 contrast ratio against their background",
        "fail_condition": "Hardcoded color pair has contrast ratio below 4.5:1 for normal text",
        "suggested_fix": "Adjust text or background color to meet 4.5:1 minimum contrast ratio",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Large text contrast is at least 3:1"): {
        "automatable": "partial",
        "inputs_required": ["html", "dom"],
        "pass_condition": "Large text (18pt+ or 14pt+ bold) color pairs meet 3:1 contrast ratio",
        "fail_condition": "Large text color pair has contrast ratio below 3:1",
        "suggested_fix": "Adjust text or background color to meet 3:1 minimum contrast ratio for large text",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Text can be resized to 200% without loss"): {
        "automatable": "no",
        "inputs_required": ["dom", "screenshot"],
        "pass_condition": "Content remains usable and no information is lost at 200% text zoom",
        "fail_condition": "Content is clipped, overlaps, or becomes unusable at 200% zoom",
        "suggested_fix": "Use relative units (rem, em, %) instead of fixed px for text and container sizes",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Content reflows without two-dimensional "): {
        "automatable": "no",
        "inputs_required": ["dom", "screenshot"],
        "pass_condition": "Content reflows to single column at 320px width without horizontal scrolling",
        "fail_condition": "Horizontal scrollbar appears at 320px viewport width for non-exempt content",
        "suggested_fix": "Use responsive design with flexible layouts, avoid fixed widths exceeding 320px",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Images of text are avoided unless essent"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "No <img> elements containing text that could be rendered as actual HTML text",
        "fail_condition": "Images are used for text content that could be implemented as styled HTML text",
        "suggested_fix": "Replace images of text with styled HTML text using CSS for visual presentation",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Text spacing changes do not break conten"): {
        "automatable": "no",
        "inputs_required": ["dom", "screenshot"],
        "pass_condition": "Content remains readable with increased line-height, letter-spacing, word-spacing, and paragraph spacing",
        "fail_condition": "Content is clipped or overlaps when text spacing is increased per WCAG 1.4.12 values",
        "suggested_fix": "Avoid fixed container heights; use flexible layouts that accommodate text spacing changes",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Normal text contrast is at least 7:1"): {
        "automatable": "partial",
        "inputs_required": ["html", "dom"],
        "pass_condition": "Hardcoded text colors meet 7:1 contrast ratio against their background",
        "fail_condition": "Hardcoded color pair has contrast ratio below 7:1 for normal text",
        "suggested_fix": "Adjust text or background color to meet 7:1 enhanced contrast ratio",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Text and visual presentation", "Large text contrast is at least 4.5:1"): {
        "automatable": "partial",
        "inputs_required": ["html", "dom"],
        "pass_condition": "Large text color pairs meet 4.5:1 contrast ratio (AAA enhanced)",
        "fail_condition": "Large text color pair has contrast ratio below 4.5:1 (AAA)",
        "suggested_fix": "Adjust text or background color to meet 4.5:1 enhanced contrast ratio for large text",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Images, icons, and tables", "Informative images have appropriate alt "): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "All <img> elements have an alt attribute; informative images have non-empty alt text",
        "fail_condition": "<img> element is missing the alt attribute entirely",
        "suggested_fix": "Add alt attribute: alt=\"description\" for informative images, alt=\"\" for decorative",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Images, icons, and tables", "Decorative images are hidden from assist"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "Decorative images have alt=\"\", role=\"presentation\", or aria-hidden=\"true\"",
        "fail_condition": "Decorative image has non-empty alt text or no alt attribute",
        "suggested_fix": "Add alt=\"\" to decorative images, or use role=\"presentation\" / aria-hidden=\"true\"",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Images, icons, and tables", "Functional images describe the action or"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Images inside links/buttons have alt text describing the action or destination, not the image appearance",
        "fail_condition": "Functional image (inside <a> or <button>) has no alt text or alt describes appearance instead of function",
        "suggested_fix": "Set alt text to describe the action (e.g., 'Search', 'Go to homepage') not the image appearance",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Images, icons, and tables", "Complex graphics have an equivalent text"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Complex images (charts, diagrams) have extended description via aria-describedby, longdesc, or adjacent text",
        "fail_condition": "Complex graphic has only a brief alt with no extended explanation available",
        "suggested_fix": "Add a detailed text description near the graphic or link to a long description",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Images, icons, and tables", "Data tables use proper table markup and "): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "<table> elements use <th> for headers, with scope or headers attributes for complex tables",
        "fail_condition": "<table> has no <th> elements, or complex table lacks scope/headers attributes",
        "suggested_fix": "Add <th> elements for column/row headers with scope=\"col\" or scope=\"row\"",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Images, icons, and tables", "Non-text UI elements and meaningful grap"): {
        "automatable": "partial",
        "inputs_required": ["html", "dom"],
        "pass_condition": "Non-text UI elements (icons, borders, focus indicators) have 3:1 contrast against adjacent colors",
        "fail_condition": "Non-text UI element has contrast ratio below 3:1 against adjacent color",
        "suggested_fix": "Increase contrast of UI element boundaries, icons, or indicators to at least 3:1",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Every form control has a programmaticall"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "Every <input>, <select>, <textarea> has an associated <label> (for/id), aria-label, or aria-labelledby",
        "fail_condition": "Form control exists without any label association (<label>, aria-label, or aria-labelledby)",
        "suggested_fix": "Add <label for=\"fieldId\"> matching the control's id, or add aria-label attribute",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Required fields are identified clearly"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Required fields have required attribute, aria-required=\"true\", or visible required indicator with text",
        "fail_condition": "Required field has no programmatic or visible required indication",
        "suggested_fix": "Add required attribute or aria-required=\"true\", plus visible text indicator",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Placeholder text is not the only label"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "Form controls with placeholder also have a visible <label>, aria-label, or aria-labelledby",
        "fail_condition": "Form control uses placeholder as the only label with no <label>, aria-label, or aria-labelledby",
        "suggested_fix": "Add a visible <label> element in addition to the placeholder attribute",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Instructions are provided before users n"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Form instructions appear before the controls they describe in DOM order",
        "fail_condition": "Instructions for form completion appear after the relevant controls or are absent",
        "suggested_fix": "Place instructions before the form or at the top of the form group",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Errors are described in text and not by "): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Error states include text messages, not just color changes (red borders, etc.)",
        "fail_condition": "Error indication uses only color (e.g., red border) without text message",
        "suggested_fix": "Add visible error text message alongside any color-based error indication",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Users can identify which field has an er"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Error messages reference or are visually adjacent to the specific field with the error",
        "fail_condition": "Error message is generic or displayed far from the erroneous field",
        "suggested_fix": "Display error messages adjacent to or within the field group, identifying the specific field",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Error messages are associated with the r"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "Error message elements are linked to their field via aria-describedby or aria-errormessage",
        "fail_condition": "Error message exists near a field but has no programmatic association (aria-describedby/aria-errormessage)",
        "suggested_fix": "Add aria-describedby on the input pointing to the error message element's id",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Input purpose is identified programmatic"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "Personal data inputs (name, email, phone, address, etc.) have appropriate autocomplete attribute",
        "fail_condition": "Personal data input field is missing autocomplete attribute",
        "suggested_fix": "Add autocomplete attribute (e.g., autocomplete=\"email\", autocomplete=\"given-name\")",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Authentication does not rely only on a c"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Authentication provides alternatives to cognitive tests (password managers, passkeys, copy-paste)",
        "fail_condition": "Authentication requires memorization or transcription with no accessible alternative",
        "suggested_fix": "Support password managers (no paste blocking), provide passkey/SSO options",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Important submissions support review, co"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Legal/financial submissions allow review, confirmation, and correction before finalizing",
        "fail_condition": "Important submission is processed immediately without review/confirm step",
        "suggested_fix": "Add a review/confirm step before final submission of important data",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Forms and validation", "Error suggestions are provided where pos"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Validation errors include suggestions for correction when the system can determine them",
        "fail_condition": "Error message states the problem but provides no guidance on how to fix it",
        "suggested_fix": "Include correction suggestions in error messages (e.g., 'Please enter a valid email like user@example.com')",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Menus, tabs, accordions, disclosures, an"): {
        "automatable": "partial",
        "inputs_required": ["html", "keyboard"],
        "pass_condition": "Composite widgets have keyboard event handlers for arrow keys, Enter, Space, Escape as appropriate",
        "fail_condition": "Interactive widget has no keyboard event handlers beyond default tab behavior",
        "suggested_fix": "Add keyboard event handlers following WAI-ARIA Authoring Practices for the widget pattern",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Expanded, collapsed, selected, active, a"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "State attributes (aria-expanded, aria-selected, aria-disabled, aria-pressed) are present on stateful components",
        "fail_condition": "Stateful component (accordion, tab, toggle) missing aria-expanded, aria-selected, or similar state attribute",
        "suggested_fix": "Add appropriate ARIA state attributes (aria-expanded, aria-selected, aria-disabled) to the control",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Dialogs have an accessible name"): {
        "automatable": "yes",
        "inputs_required": ["html"],
        "pass_condition": "Dialog elements have aria-label or aria-labelledby pointing to a visible heading",
        "fail_condition": "<dialog> or role=\"dialog\" element has no aria-label or aria-labelledby",
        "suggested_fix": "Add aria-labelledby pointing to the dialog's heading element, or add aria-label",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Focus moves into dialogs when they open"): {
        "automatable": "partial",
        "inputs_required": ["html", "keyboard"],
        "pass_condition": "Dialog open logic includes focus management (element.focus() or autofocus attribute)",
        "fail_condition": "Dialog opens without any focus management code",
        "suggested_fix": "Call .focus() on the dialog or its first focusable element when opening, or use autofocus attribute",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Focus is contained within modal dialogs "): {
        "automatable": "partial",
        "inputs_required": ["html", "keyboard"],
        "pass_condition": "Modal dialog implements focus trapping (Tab cycles within dialog) or uses native <dialog> with showModal()",
        "fail_condition": "Modal dialog allows Tab to move focus outside the dialog while it is open",
        "suggested_fix": "Use native <dialog>.showModal() which traps focus, or implement a focus trap with keydown listener",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Focus returns to a logical place when di"): {
        "automatable": "partial",
        "inputs_required": ["html", "keyboard"],
        "pass_condition": "Dialog close logic restores focus to the triggering element",
        "fail_condition": "Dialog closes without restoring focus to the trigger or another logical element",
        "suggested_fix": "Store a reference to the trigger element before opening, and call trigger.focus() on close",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Background content is not interactable w"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Modal uses <dialog>.showModal(), inert attribute on background, or aria-hidden on background content",
        "fail_condition": "Modal is open but background content remains interactive (no inert, no aria-hidden on background)",
        "suggested_fix": "Use <dialog>.showModal() or add inert attribute to background content while modal is open",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Important dynamic updates are announced "): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Dynamic content containers have aria-live, role=\"alert\", role=\"status\", or role=\"log\"",
        "fail_condition": "Dynamic content updates without any live region markup",
        "suggested_fix": "Add aria-live=\"polite\" (or role=\"status\") to the container that receives dynamic updates",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Status, success, warning, and error mess"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Status message containers use role=\"status\", role=\"alert\", or aria-live with appropriate politeness",
        "fail_condition": "Status/toast message appears in DOM without live region attributes",
        "suggested_fix": "Wrap status messages in an element with role=\"status\" (polite) or role=\"alert\" (assertive)",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Components and dynamic UI", "Dynamic updates do not steal focus unles"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Dynamic content updates use aria-live instead of programmatic focus movement",
        "fail_condition": "Non-critical dynamic update moves focus away from user's current position",
        "suggested_fix": "Use aria-live regions for non-critical updates instead of moving focus",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Prerecorded audio has a transcript or eq"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "<audio> elements have an adjacent or linked transcript",
        "fail_condition": "<audio> element exists with no associated transcript link or adjacent text",
        "suggested_fix": "Provide a transcript link near the audio player or include transcript text on the page",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Prerecorded video with audio has caption"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "<video> elements have a <track kind=\"captions\"> element",
        "fail_condition": "<video> element has no <track kind=\"captions\"> child",
        "suggested_fix": "Add <track kind=\"captions\" src=\"captions.vtt\" srclang=\"en\" label=\"English\"> to the video element",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Video-only content has an equivalent tex"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Video-only elements have an adjacent or linked text alternative",
        "fail_condition": "Video-only content has no text or audio alternative",
        "suggested_fix": "Provide a text description or audio alternative for video-only content",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Auto-playing audio longer than 3 seconds"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "<audio>/<video> with autoplay has controls attribute or a visible pause/stop mechanism",
        "fail_condition": "Media element has autoplay without controls attribute or visible pause mechanism",
        "suggested_fix": "Add controls attribute to the media element, or provide a visible pause/stop button",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Moving or auto-updating content can be p"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Animated/auto-updating elements have a visible pause, stop, or hide mechanism",
        "fail_condition": "Auto-moving/updating content (carousel, ticker) has no pause/stop control",
        "suggested_fix": "Add a pause/stop button for carousels, tickers, and auto-updating content",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Content does not flash above the allowed"): {
        "automatable": "no",
        "inputs_required": ["screenshot", "media"],
        "pass_condition": "No content flashes more than 3 times per second",
        "fail_condition": "Content flashes more than 3 times per second or exceeds general flash thresholds",
        "suggested_fix": "Reduce flash frequency below 3 per second, or reduce the flashing area",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Time limits can be turned off, adjusted,"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Any time limit provides turn off, adjust, or extend mechanism before expiry",
        "fail_condition": "Session or task timeout occurs without warning or extension option",
        "suggested_fix": "Add a timeout warning with option to extend, or allow users to disable time limits",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Dragging interactions have a simple poin"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Elements with drag handlers also have click/tap alternative or button-based controls",
        "fail_condition": "Element uses drag events (dragstart, ondrag) without a simple pointer alternative",
        "suggested_fix": "Add button-based or click alternative for drag interactions (e.g., move up/down buttons)",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Pointer actions can be canceled where ap"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Actions fire on mouseup/pointerup/click (not mousedown/pointerdown) allowing cancel by moving away",
        "fail_condition": "Actions fire on mousedown/pointerdown without cancel mechanism",
        "suggested_fix": "Move action handlers from mousedown/pointerdown to click/mouseup/pointerup events",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Prerecorded video includes audio descrip"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "<video> has <track kind=\"descriptions\"> or a linked audio description alternative",
        "fail_condition": "Video with meaningful visual content has no audio description track or alternative",
        "suggested_fix": "Add <track kind=\"descriptions\"> or provide a separate audio-described version",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Live media has captions where required"): {
        "automatable": "no",
        "inputs_required": ["html", "media"],
        "pass_condition": "Live media streams include real-time captions",
        "fail_condition": "Live media stream has no captioning mechanism",
        "suggested_fix": "Integrate real-time captioning service for live media",
        "severity": "major",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Media, motion, and timing", "Motion-based interaction can be disabled"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Motion-based features (shake, tilt) can be disabled and have button/click alternatives",
        "fail_condition": "Functionality requires device motion with no setting to disable or alternative input",
        "suggested_fix": "Provide a setting to disable motion-based interaction and offer button/click alternatives",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Consistency and compatibility", "Navigation patterns are used consistentl"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Navigation components maintain the same order and structure across pages",
        "fail_condition": "Navigation order or structure varies between pages",
        "suggested_fix": "Use shared navigation components/templates to ensure consistency across pages",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Consistency and compatibility", "Components with the same function are id"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Components serving the same function use the same label, icon, and text across all occurrences",
        "fail_condition": "Same-function components use different labels or icons on different pages",
        "suggested_fix": "Standardize labels and icons for recurring components (e.g., search, print, save)",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Consistency and compatibility", "Accessibility APIs are updated correctly"): {
        "automatable": "partial",
        "inputs_required": ["html"],
        "pass_condition": "Custom components update ARIA attributes when state changes (aria-expanded, aria-checked, etc.)",
        "fail_condition": "Component changes visual state but does not update corresponding ARIA attributes",
        "suggested_fix": "Update ARIA state attributes programmatically when component state changes",
        "severity": "critical",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    },
    ("Consistency and compatibility", "Accessibility is verified in the browser"): {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Testing matrix includes browser + AT combinations the product supports",
        "fail_condition": "No documented browser/AT testing matrix or test evidence",
        "suggested_fix": "Define and document a browser/AT test matrix; include in QA process",
        "severity": "minor",
        "file_patterns": [],
    },
}


def get_enrichment(category: str, item: str) -> dict:
    """Look up enrichment data by matching category + first 40 chars of item."""
    key = (category, item[:40])
    default = {
        "automatable": "no",
        "inputs_required": ["html"],
        "pass_condition": "Manual verification required",
        "fail_condition": "Manual verification required",
        "suggested_fix": "Review manually per the manual check instruction",
        "severity": "minor",
        "file_patterns": ["*.html", "*.vue", "*.jsx", "*.tsx"],
    }
    return ENRICHMENTS.get(key, default)


def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    project_root = os.path.dirname(script_dir)
    output_dir = os.path.join(project_root, "checklist")
    xlsx_path = os.path.join(output_dir, "accessibility_checklist_audit.xlsx")
    output_path = os.path.join(output_dir, "wcag22-checklist.json")

    if not os.path.exists(xlsx_path):
        print(f"Error: {xlsx_path} not found", file=sys.stderr)
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb["Master Checklist"]

    checks = []
    counters = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        category = row[0]
        item = row[1]
        level = row[2]

        if not category or not item:
            continue

        if level == "Process":
            continue

        applies_to = row[6] or "All"
        manual_check = row[7] or ""
        wcag_sc = row[8] or ""
        sc_name = row[9] or ""

        enrichment = get_enrichment(category, item)

        check = {
            "id": make_id(category, item, counters),
            "category": category,
            "item": item,
            "level": level,
            "wcag_sc": parse_wcag_sc(wcag_sc),
            "sc_name": sc_name,
            "applies_to": parse_applies_to(applies_to),
            "manual_check": manual_check,
            "automatable": enrichment["automatable"],
            "inputs_required": enrichment["inputs_required"],
            "pass_condition": enrichment["pass_condition"],
            "fail_condition": enrichment["fail_condition"],
            "suggested_fix": enrichment["suggested_fix"],
            "severity": enrichment["severity"],
            "file_patterns": enrichment["file_patterns"],
        }
        checks.append(check)

    wb.close()

    output = {
        "version": "1.0.0",
        "wcag_version": "2.2",
        "generated_from": "accessibility_checklist_audit.xlsx",
        "total_checks": len(checks),
        "checks": checks,
    }

    with open(output_path, "w") as f:
        json.dump(output, f, indent=2)

    print(f"Exported {len(checks)} checks to {output_path}")

    auto_yes = sum(1 for c in checks if c["automatable"] == "yes")
    auto_partial = sum(1 for c in checks if c["automatable"] == "partial")
    auto_no = sum(1 for c in checks if c["automatable"] == "no")
    print(f"  Fully automatable: {auto_yes}")
    print(f"  Partially automatable: {auto_partial}")
    print(f"  Manual only: {auto_no}")


if __name__ == "__main__":
    main()
